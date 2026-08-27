from __future__ import annotations

import csv
import io
import re
import unicodedata
from typing import Any

from blackbeans_api.leads.scoring import LEAD_QUALITY_FIELDS
from blackbeans_api.leads.scoring import apply_prospect_quality
from blackbeans_api.leads.scoring import compute_prospect_score
from blackbeans_api.leads.scoring import extract_job_title
from blackbeans_api.leads.scoring import is_valid_cnpj
from blackbeans_api.leads.scoring import prospect_quality_from_lead

DISPLAY_NAME_KEYS = (
    "nome",
    "nome da empresa",
    "nome empresa",
    "cliente",
    "razao_social_rf",
    "razao social",
    "razão social",
    "contato",
    "nome contato",
    "nome responsavel",
    "nome responsável",
    "company",
    "name",
)

COMPANY_NAME_KEYS = (
    "nome da empresa",
    "nome empresa",
    "razao_social_rf",
    "razao social",
    "razão social",
    "razao_social",
    "cliente",
    "company",
    "empresa",
    "construtora",
    "incorporadora",
)

CONTACT_NAME_KEYS = (
    "nome contato",
    "nome do contato",
    "nome responsavel",
    "nome responsável",
    "contato",
    "nome",
    "name",
)

CNPJ_KEY_HINTS = ("cnpj",)
EMAIL_KEY_HINTS = ("email", "e-mail", "mail")
PHONE_KEY_HINTS = ("telefone", "phone", "celular", "whatsapp", "tel", "fone")
SITE_KEY_HINTS = ("site", "website", "url", "web")
ADDRESS_KEY_HINTS = (
    "endereco",
    "endereço",
    "logradouro",
    "rua",
    "bairro",
    "cidade",
    "cep",
)

_EMAIL_RE = re.compile(r"[^@\s]+@[^@\s]+\.[^@\s]+")
_CNPJ_DIGITS_RE = re.compile(r"\D+")


class LeadParseError(Exception):
    """Arquivo invalido ou sem dados utilizaveis."""


def _cell_to_str(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float):
        if value.is_integer():
            text = str(int(value))
        else:
            text = str(value)
    else:
        text = str(value).strip()
    if not text or text.lower() in {"none", "null", "nan"}:
        return None
    return text


def _normalize_header(raw: Any, index: int) -> str:
    text = _cell_to_str(raw)
    if not text:
        return f"coluna_{index + 1}"
    return text


def normalize_company_name(name: str | None) -> str:
    if not name:
        return ""
    text = unicodedata.normalize("NFKD", name)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-zA-Z0-9\s]", " ", text.lower())
    text = re.sub(r"\s+", " ", text).strip()
    return text[:512]


def normalize_cnpj(value: str | None) -> str | None:
    if not value:
        return None
    digits = _CNPJ_DIGITS_RE.sub("", value)
    if len(digits) != 14:
        return None
    return digits


def normalize_phone(value: str | None) -> str | None:
    if not value:
        return None
    text = value.strip()
    if text.lower().startswith("tel:"):
        text = text[4:]
    digits = re.sub(r"\D", "", text)
    if len(digits) < 8:
        return None
    return digits[:20]


def normalize_email(value: str | None) -> str | None:
    if not value:
        return None
    match = _EMAIL_RE.search(value.strip())
    if not match:
        return None
    return match.group(0).lower()[:255]


def _key_map(payload: dict[str, Any]) -> dict[str, str]:
    return {str(k).strip().lower(): k for k in payload}


def _first_by_keys(payload: dict[str, Any], candidates: tuple[str, ...]) -> str | None:
    key_map = _key_map(payload)
    for candidate in candidates:
        real = key_map.get(candidate)
        if real is None:
            continue
        value = _cell_to_str(payload.get(real))
        if value:
            return value
    return None


def _first_by_hint(payload: dict[str, Any], hints: tuple[str, ...]) -> str | None:
    for key, raw in payload.items():
        low = str(key).strip().lower()
        if any(h in low for h in hints):
            value = _cell_to_str(raw)
            if value:
                return value
    return None


def derive_display_name(payload: dict[str, Any], column_keys: list[str]) -> str:
    key_map = _key_map(payload)
    for candidate in DISPLAY_NAME_KEYS:
        real = key_map.get(candidate)
        if real is None:
            continue
        value = _cell_to_str(payload.get(real))
        if value:
            return value[:512]
    for key in column_keys:
        value = _cell_to_str(payload.get(key))
        if value:
            return value[:512]
    for value in payload.values():
        text = _cell_to_str(value)
        if text:
            return text[:512]
    return "Lead sem nome"


def derive_company_name(payload: dict[str, Any], *, fallback: str | None = None) -> str:
    value = _first_by_keys(payload, COMPANY_NAME_KEYS)
    if value:
        return value[:512]
    if fallback:
        return fallback[:512]
    return "Empresa sem nome"


def derive_contact_name(payload: dict[str, Any], *, fallback: str | None = None) -> str:
    value = _first_by_keys(payload, CONTACT_NAME_KEYS)
    if value:
        return value[:512]
    if fallback:
        return fallback[:512]
    return "Contato sem nome"


def extract_structured_fields(payload: dict[str, Any]) -> dict[str, Any]:
    cnpj_raw = _first_by_hint(payload, CNPJ_KEY_HINTS) or _first_by_keys(
        payload, ("cnpj",)
    )
    email_raw = _first_by_hint(payload, EMAIL_KEY_HINTS)
    phone_raw = _first_by_hint(payload, PHONE_KEY_HINTS)
    return {
        "cnpj": normalize_cnpj(cnpj_raw),
        "email": normalize_email(email_raw),
        "phone": normalize_phone(phone_raw),
        "has_site": bool(_first_by_hint(payload, SITE_KEY_HINTS)),
        "has_address": bool(_first_by_hint(payload, ADDRESS_KEY_HINTS)),
        "job_title": extract_job_title(payload),
    }


def enrich_lead_fields(
    payload: dict[str, Any], column_keys: list[str]
) -> dict[str, Any]:
    display_name = derive_display_name(payload, column_keys)
    contact_name = derive_contact_name(payload, fallback=display_name)
    company_name = derive_company_name(payload, fallback=display_name)
    structured = extract_structured_fields(payload)
    cnpj = structured["cnpj"] if isinstance(structured["cnpj"], str) else None
    email = structured["email"] if isinstance(structured["email"], str) else None
    phone = structured["phone"] if isinstance(structured["phone"], str) else None
    quality = compute_prospect_score(
        cnpj=cnpj,
        email=email,
        phone=phone,
        contact_name=contact_name,
        company_name=company_name,
        job_title=str(structured.get("job_title") or ""),
    )
    return {
        "display_name": contact_name,
        "company_name": company_name,
        "company_name_normalized": normalize_company_name(company_name),
        "cnpj": cnpj or "",
        "email": email or "",
        "phone": phone or "",
        "job_title": structured.get("job_title") or "",
        "has_cnpj": quality["has_cnpj"],
        "has_email": quality["has_email"],
        "has_phone": quality["has_phone"],
        "has_site": bool(structured.get("has_site")),
        "has_address": bool(structured.get("has_address")),
        "completeness_score": quality["completeness_score"],
        "email_is_generic": quality["email_is_generic"],
        "email_is_shared": quality["email_is_shared"],
        "phone_is_shared": quality["phone_is_shared"],
        "contact_is_person": quality["contact_is_person"],
        "contact_is_decision_maker": quality["contact_is_decision_maker"],
        "score_breakdown": quality["score_breakdown"],
    }


def build_search_text(
    *, payload: dict[str, Any], origem: str, display_name: str
) -> str:
    parts: list[str] = [origem or "", display_name or ""]
    for value in payload.values():
        text = _cell_to_str(value)
        if text:
            parts.append(text)
    return " ".join(parts).lower()


def build_company_search_text(
    *, name: str, cnpj: str | None, origem: str, extra: str = ""
) -> str:
    return " ".join(part for part in [name, cnpj or "", origem, extra] if part).lower()


def recompute_company_quality(company) -> None:
    """Atualiza flags/score/contagem da empresa a partir da qualidade dos contatos."""
    contacts = list(company.contacts.all())
    emails = [row for row in contacts if row.email]
    phones = [row for row in contacts if row.phone]
    company.has_cnpj = is_valid_cnpj(company.cnpj) or any(
        row.has_cnpj for row in contacts
    )
    company.has_phone = any(row.has_phone for row in contacts)
    company.has_email = any(row.has_email for row in contacts)
    company.completeness_score = max(
        (row.completeness_score for row in contacts), default=0
    )
    company.email_is_generic = bool(emails) and all(
        row.email_is_generic for row in emails
    )
    company.email_is_shared = bool(emails) and all(
        row.email_is_shared for row in emails
    )
    company.phone_is_shared = bool(phones) and all(
        row.phone_is_shared for row in phones
    )
    company.contact_is_person = any(row.contact_is_person for row in contacts)
    company.contact_is_decision_maker = any(
        row.contact_is_decision_maker for row in contacts
    )
    company.contacts_count = len(contacts)
    company.search_text = build_company_search_text(
        name=company.name,
        cnpj=company.cnpj,
        origem=company.origem,
        extra=" ".join(c.display_name for c in contacts[:20]),
    )
    company.save(
        update_fields=[
            "has_cnpj",
            "has_phone",
            "has_email",
            "completeness_score",
            "email_is_generic",
            "email_is_shared",
            "phone_is_shared",
            "contact_is_person",
            "contact_is_decision_maker",
            "contacts_count",
            "search_text",
            "updated_at",
        ],
    )


def _shared_values(field: str, values: list[str] | None = None) -> set[str]:
    from django.db.models import Count

    from blackbeans_api.leads.models import Lead

    queryset = Lead.objects.exclude(**{field: ""})
    if values is not None:
        cleaned = [item for item in values if item]
        if not cleaned:
            return set()
        queryset = queryset.filter(**{f"{field}__in": cleaned})
    return set(
        queryset.values(field)
        .annotate(n=Count("id"))
        .filter(n__gte=2)
        .values_list(field, flat=True),
    )


def _leads_affected_by_contacts(
    *,
    emails: list[str] | None,
    phones: list[str] | None,
):
    from django.db.models import Q

    from blackbeans_api.leads.models import Lead

    cleaned_emails = [item for item in (emails or []) if item]
    cleaned_phones = [item for item in (phones or []) if item]
    clause = Q()
    if cleaned_emails:
        clause |= Q(email__in=cleaned_emails)
    if cleaned_phones:
        clause |= Q(phone__in=cleaned_phones)
    if not clause:
        return Lead.objects.none()
    seed = Lead.objects.filter(clause)
    related_emails = set(seed.exclude(email="").values_list("email", flat=True))
    related_phones = set(seed.exclude(phone="").values_list("phone", flat=True))
    related_emails.update(cleaned_emails)
    related_phones.update(cleaned_phones)
    expand = Q()
    if related_emails:
        expand |= Q(email__in=related_emails)
    if related_phones:
        expand |= Q(phone__in=related_phones)
    return Lead.objects.select_related("company").filter(expand)


def refresh_shared_quality(
    *,
    emails: list[str] | None = None,
    phones: list[str] | None = None,
    batch_size: int = 500,
) -> dict[str, int]:
    """Recalcula flags de duplicata e o score dos leads afetados (e das empresas)."""
    from django.utils import timezone

    from blackbeans_api.leads.models import Lead
    from blackbeans_api.leads.models import LeadCompany

    if emails is not None or phones is not None:
        queryset = _leads_affected_by_contacts(emails=emails, phones=phones)
    else:
        queryset = Lead.objects.select_related("company").all()

    leads = list(queryset)
    if not leads:
        return {"leads": 0, "companies": 0}

    shared_emails = _shared_values("email", [row.email for row in leads if row.email])
    shared_phones = _shared_values("phone", [row.phone for row in leads if row.phone])
    now = timezone.now()
    company_ids: set[str] = set()
    for lead in leads:
        quality = prospect_quality_from_lead(
            lead,
            email_is_shared=bool(lead.email) and lead.email in shared_emails,
            phone_is_shared=bool(lead.phone) and lead.phone in shared_phones,
        )
        apply_prospect_quality(lead, quality)
        lead.updated_at = now
        if lead.company_id:
            company_ids.add(str(lead.company_id))

    Lead.objects.bulk_update(
        leads,
        list(LEAD_QUALITY_FIELDS) + ["updated_at"],
        batch_size=batch_size,
    )
    if company_ids:
        for company in LeadCompany.objects.filter(pk__in=company_ids).prefetch_related(
            "contacts",
        ):
            recompute_company_quality(company)
    return {"leads": len(leads), "companies": len(company_ids)}


def recompute_all_lead_scores(*, batch_size: int = 500) -> dict[str, int]:
    """Recalcula o score de prospecção de todos os leads e empresas."""
    return refresh_shared_quality(batch_size=batch_size)


def get_or_create_company_for_payload(
    *,
    payload: dict[str, Any],
    column_keys: list[str],
    origem: str,
    freshness: str,
    cache: dict[str, Any] | None = None,
):
    from blackbeans_api.leads.models import LeadCompany

    enriched = enrich_lead_fields(payload, column_keys)
    cnpj = enriched["cnpj"] or None
    name = enriched["company_name"]
    name_norm = enriched["company_name_normalized"]
    cache = cache if cache is not None else {}

    cache_key = f"cnpj:{cnpj}" if cnpj else f"name:{name_norm}"
    if cache_key in cache:
        company = cache[cache_key]
    else:
        company = None
        if cnpj:
            company = LeadCompany.objects.filter(cnpj=cnpj).first()
        if company is None and name_norm:
            company = LeadCompany.objects.filter(name_normalized=name_norm).first()
        if company is None:
            company = LeadCompany.objects.create(
                name=name,
                name_normalized=name_norm,
                cnpj=cnpj,
                origem=origem or "",
                freshness=freshness,
                has_cnpj=bool(cnpj) and is_valid_cnpj(cnpj),
                has_phone=enriched["has_phone"],
                has_email=enriched["has_email"],
                completeness_score=enriched["completeness_score"],
                email_is_generic=enriched["email_is_generic"],
                email_is_shared=enriched["email_is_shared"],
                phone_is_shared=enriched["phone_is_shared"],
                contact_is_person=enriched["contact_is_person"],
                contact_is_decision_maker=enriched["contact_is_decision_maker"],
                contacts_count=0,
                search_text=build_company_search_text(
                    name=name, cnpj=cnpj, origem=origem
                ),
            )
        else:
            changed = False
            if cnpj and not company.cnpj:
                company.cnpj = cnpj
                company.has_cnpj = is_valid_cnpj(cnpj)
                changed = True
            if origem and not company.origem:
                company.origem = origem
                changed = True
            if freshness == "novo" and company.freshness != "novo":
                company.freshness = "novo"
                changed = True
            if changed:
                company.save()
        cache[cache_key] = company
    return company, enriched


_LEAD_COMPANY_BACKFILL_LOCK = "leads:backfill_companies"
_LEAD_BACKFILL_FIELDS = (
    "company",
    "display_name",
    "email",
    "phone",
    "cnpj",
    "has_cnpj",
    "has_phone",
    "has_email",
    "completeness_score",
    "email_is_generic",
    "email_is_shared",
    "phone_is_shared",
    "contact_is_person",
    "contact_is_decision_maker",
    "search_text",
    "updated_at",
)


def backfill_lead_companies(  # noqa: C901
    *,
    only_missing: bool = True,
    batch_size: int = 200,
) -> dict[str, int]:
    """Associa leads a LeadCompany e recalcula qualidade."""
    from django.db import transaction
    from django.db.models import Count
    from django.utils import timezone

    from blackbeans_api.leads.models import Lead
    from blackbeans_api.leads.models import LeadCompany

    batch_size = max(1, int(batch_size))
    processed = 0
    company_cache: dict[str, Any] = {}
    touched: dict[str, Any] = {}
    last_pk = None

    while True:
        queryset = Lead.objects.select_related("import_batch", "company").order_by("pk")
        if only_missing:
            queryset = queryset.filter(company__isnull=True)
        if last_pk is not None:
            queryset = queryset.filter(pk__gt=last_pk)
        batch = list(queryset[:batch_size])
        if not batch:
            break
        last_pk = batch[-1].pk
        now = timezone.now()
        with transaction.atomic():
            for lead in batch:
                column_keys = list(
                    (lead.import_batch.column_keys if lead.import_batch else None) or []
                )
                if not column_keys:
                    column_keys = list((lead.payload or {}).keys())
                origem = ""
                freshness = "novo"
                if lead.import_batch:
                    origem = lead.import_batch.origem
                    freshness = lead.import_batch.freshness
                company, enriched = get_or_create_company_for_payload(
                    payload=dict(lead.payload or {}),
                    column_keys=column_keys,
                    origem=origem,
                    freshness=freshness,
                    cache=company_cache,
                )
                lead.company = company
                lead.display_name = enriched["display_name"] or lead.display_name
                lead.email = enriched["email"]
                lead.phone = enriched["phone"]
                lead.cnpj = enriched["cnpj"]
                lead.has_cnpj = enriched["has_cnpj"]
                lead.has_phone = enriched["has_phone"]
                lead.has_email = enriched["has_email"]
                lead.completeness_score = enriched["completeness_score"]
                lead.email_is_generic = enriched["email_is_generic"]
                lead.email_is_shared = enriched["email_is_shared"]
                lead.phone_is_shared = enriched["phone_is_shared"]
                lead.contact_is_person = enriched["contact_is_person"]
                lead.contact_is_decision_maker = enriched["contact_is_decision_maker"]
                lead.search_text = build_search_text(
                    payload=dict(lead.payload or {}),
                    origem=origem,
                    display_name=lead.display_name,
                )
                lead.updated_at = now
                touched[str(company.pk)] = company
                processed += 1
            Lead.objects.bulk_update(
                batch, list(_LEAD_BACKFILL_FIELDS), batch_size=batch_size
            )

    company_ids = list(touched.keys())
    if company_ids:
        companies = LeadCompany.objects.filter(pk__in=company_ids).prefetch_related(
            "contacts"
        )
        for company in companies:
            recompute_company_quality(company)

    refresh_shared_quality()

    orphan_companies = LeadCompany.objects.annotate(n=Count("contacts")).filter(n=0)
    orphan_count = 0
    for company in orphan_companies.iterator():
        recompute_company_quality(company)
        orphan_count += 1

    return {
        "processed": processed,
        "companies": len(company_ids),
        "orphans": orphan_count,
    }


def ensure_orphan_leads_have_companies() -> dict[str, int] | None:
    """Roda o backfill se ainda existirem leads sem empresa."""
    from django.core.cache import cache

    from blackbeans_api.leads.models import Lead

    if not Lead.objects.filter(company__isnull=True).exists():
        return None
    if not cache.add(_LEAD_COMPANY_BACKFILL_LOCK, "1", timeout=900):
        return None
    try:
        return backfill_lead_companies(only_missing=True)
    finally:
        cache.delete(_LEAD_COMPANY_BACKFILL_LOCK)


def _rows_from_matrix(
    matrix: list[list[Any]],
) -> tuple[list[str], list[dict[str, Any]]]:
    if not matrix:
        raise LeadParseError("Planilha vazia.")
    header_row = matrix[0]
    if not header_row:
        raise LeadParseError("Cabecalho da planilha vazio.")

    column_keys: list[str] = []
    seen: dict[str, int] = {}
    for idx, raw in enumerate(header_row):
        base = _normalize_header(raw, idx)
        count = seen.get(base, 0)
        seen[base] = count + 1
        column_keys.append(base if count == 0 else f"{base}_{count + 1}")

    rows: list[dict[str, Any]] = []
    for raw_row in matrix[1:]:
        if raw_row is None:
            continue
        payload: dict[str, Any] = {}
        has_value = False
        for idx, key in enumerate(column_keys):
            cell = raw_row[idx] if idx < len(raw_row) else None
            value = _cell_to_str(cell)
            if value is not None:
                payload[key] = value
                has_value = True
        if has_value:
            rows.append(payload)

    if not rows:
        raise LeadParseError("Nenhuma linha de dados encontrada.")
    return column_keys, rows


def parse_csv_bytes(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    text: str | None = None
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise LeadParseError("Nao foi possivel decodificar o CSV.")

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(io.StringIO(text), dialect)
    matrix = [list(row) for row in reader]
    return _rows_from_matrix(matrix)


def parse_xlsx_bytes(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise LeadParseError("Suporte a XLSX indisponivel (openpyxl).") from exc

    try:
        workbook = load_workbook(
            filename=io.BytesIO(content), read_only=True, data_only=True
        )
    except Exception as exc:
        raise LeadParseError("Arquivo XLSX invalido.") from exc

    try:
        sheet = workbook.active
        matrix: list[list[Any]] = []
        for row in sheet.iter_rows(values_only=True):
            matrix.append(list(row) if row else [])
    finally:
        workbook.close()

    return _rows_from_matrix(matrix)


def parse_spreadsheet(
    *, filename: str, content: bytes
) -> tuple[list[str], list[dict[str, Any]]]:
    name = (filename or "").lower().strip()
    if name.endswith(".csv") or name.endswith(".txt"):
        return parse_csv_bytes(content)
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return parse_xlsx_bytes(content)
    if content[:2] == b"PK":
        return parse_xlsx_bytes(content)
    return parse_csv_bytes(content)
